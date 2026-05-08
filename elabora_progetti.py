"""
elabora_progetti.py — Report settimanale risorse consulenza Red Hat Italy

Legge un file Excel di input (export da PSA/pianificazione) e produce un file
Excel di output multi-foglio con:
  - dati     : dati sorgente arricchiti con colonne derivate
  - progetti : riepilogo contratti con giorni consuntivati vs. riscattati
  - Riepilogo Settimanale : pivot actual/estimated per progetto e settimana
  - Dettaglio Ruoli       : pivot estimated con breakdown per ruolo/milestone
  - Tabella di Export     : riepilogo giornate per codice ordine

Utilizzo:
    python elabora_progetti.py [cliente] [input.xlsx] [cust.config] [output.xlsx]

    Tutti gli argomenti sono opzionali; i default sono:
        cliente='', input.xlsx, cust.config, output_elaborato.xlsx

    cliente:     filtro sulla colonna "Cliente" (case-insensitive; vuoto = tutti).
    cust.config: impostazioni specifiche del cliente (contratti, Export, contatti).

    Le impostazioni generiche vengono sempre lette da script.config (fisso).
"""

import pandas as pd
import os
import re
import sys
import logging
import subprocess
import traceback
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
import json

logging.basicConfig(level=logging.INFO, format='%(message)s')
log = logging.getLogger(__name__)

# --- FUNZIONI DI SUPPORTO ---

def carica_config(nome_file):
    """Legge un file .config chiave=valore e restituisce un dizionario.

    Ignora righe vuote e righe che iniziano con '#'.
    Usa UTF-8 per supportare caratteri accentati nei valori.

    Args:
        nome_file: percorso del file di configurazione.

    Returns:
        dict con le coppie chiave/valore; dizionario vuoto se il file non esiste.
    """
    config = {}
    if not os.path.exists(nome_file):
        return config
    with open(nome_file, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line and "=" in line and not line.startswith('#'):
                key, value = line.split('=', 1)
                config[key.strip()] = value.strip()
    return config

def indicizza_contratti(config):
    """Costruisce un indice inverso {nome_contratto -> suffisso} dalla config.

    Le chiavi ContractName1, ContractName2, … vengono mappate ai rispettivi
    suffissi numerici in modo da recuperare velocemente Opportunity, EndDate e
    DaysRedempted per ogni contratto senza scorrere tutta la config ogni volta.

    Args:
        config: dizionario restituito da carica_config.

    Returns:
        dict {contract_name_string: suffix_string}
    """
    contratti = {}
    for key, value in config.items():
        if key.startswith("ContractName"):
            suffix = key.replace("ContractName", "")
            contratti[value.strip()] = suffix
    return contratti

def trova_valore_config(contract_name, contratti_idx, config, prefisso):
    """Recupera un valore di config per un contratto dato il suo nome.

    Esempio: trova_valore_config("ISP_GPSteam25...", idx, cfg, "Opportunity")
    restituisce il valore di Opportunity1 se ContractName1=="ISP_GPSteam25...".

    Args:
        contract_name:  nome del contratto (valore colonna A del sorgente).
        contratti_idx:  indice restituito da indicizza_contratti.
        config:         dizionario della configurazione.
        prefisso:       prefisso della chiave da cercare (es. "Opportunity").

    Returns:
        Stringa con il valore trovato, oppure stringa vuota se non presente.
    """
    suffix = contratti_idx.get(str(contract_name).strip())
    if suffix is None:
        return ""
    return config.get(f"{prefisso}{suffix}", "")

def estrai_settimana(testo):
    """Estrae il numero di settimana da una stringa nel formato *W<nn>*.

    Usata per ricavare la settimana ISO dal campo periodo (es. "CY2025-W14").

    Args:
        testo: stringa o NaN.

    Returns:
        int con il numero di settimana, o None se non trovato/NaN.
    """
    if pd.isna(testo):
        return None
    match = re.search(r'W(\d+)', str(testo), re.IGNORECASE)
    return int(match.group(1)) if match else None

def get_date_range(year, start_week, end_week):
    """Calcola la data di inizio e fine di un intervallo di settimane ISO.

    Usa il 4 gennaio come ancora ISO per trovare il lunedì della prima settimana
    (conforme alla norma ISO 8601).

    Args:
        year:       anno di riferimento.
        start_week: prima settimana ISO dell'intervallo.
        end_week:   ultima settimana ISO dell'intervallo.

    Returns:
        Tuple (data_inizio, data_fine) in formato "dd/mm/yyyy"; ("N/A", "N/A")
        in caso di errore.
    """
    try:
        d = datetime(year, 1, 4)
        start_date = d + timedelta(weeks=(start_week - 1), days=-d.weekday())
        end_date = d + timedelta(weeks=(end_week - 1), days=-d.weekday() + 6)
        return start_date.strftime("%d/%m/%Y"), end_date.strftime("%d/%m/%Y")
    except Exception:
        return "N/A", "N/A"

def calcola_date_settimana(week_str):
    """Restituisce l'intervallo di date (lun-dom) per una settimana ISO.

    Args:
        week_str: stringa nel formato "YYYY-Www" (es. "2025-W14").

    Returns:
        Stringa "dd/mm - dd/mm" con inizio e fine settimana, o None se il
        formato non è riconosciuto.
    """
    match = re.search(r'(\d{4})-W(\d+)', week_str, re.IGNORECASE)
    if not match:
        return None
    year, week = int(match.group(1)), int(match.group(2))
    try:
        lunedi = datetime.fromisocalendar(year, week, 1)
        domenica = datetime.fromisocalendar(year, week, 7)
        return f"{lunedi.strftime('%d/%m')} - {domenica.strftime('%d/%m')}"
    except Exception:
        return None

def splitta_assegnazione(val):
    """Divide il campo assegnazione (colonna C del sorgente) nei suoi 7 componenti.

    Il formato atteso è: "Nome|OPA@profilo|Cliente|SottoProgetto|Riferimento|Commento"
    Il campo Riferimento può contenere due valori numerici separati da '&'
    (es. "6&4" o "6 & 4"): il primo va in "Riferimento tabella 1", il secondo
    in "Sotto Riferimento tabella 1". Con un solo valore, "Sotto Riferimento
    tabella 1" resta vuoto. I riferimenti numerici sono convertiti in interi.

    Args:
        val: stringa grezza della cella, o NaN.

    Returns:
        Lista di esattamente 7 elementi.
    """
    parti = [p.strip() for p in str(val).split('|')] if pd.notna(val) else []
    while len(parti) < 6:
        parti.append("")
    parti = parti[:6]

    def _to_int(s):
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return s

    rif_parts = [r.strip() for r in re.split(r'\s*&\s*', parti[4]) if r.strip()]
    rif1 = _to_int(rif_parts[0]) if len(rif_parts) >= 1 else ''
    rif2 = _to_int(rif_parts[1]) if len(rif_parts) >= 2 else ''

    return [parti[0], parti[1], parti[2], parti[3], rif1, rif2, parti[5]]

# --- CARICAMENTO E PREPARAZIONE DATI ---

def carica_dati(file_excel_input, config):
    """Carica il file Excel sorgente e prepara il DataFrame arricchito.

    Le colonne di interesse (ruolo, ore stimate, ore effettive, progetto, periodo)
    vengono lette per indice: i valori di default sono compatibili con il formato
    PSA standard ma possono essere sovrascritta tramite config (ColIdxRuolo, ecc.).

    Espande anche la colonna assegnazione (colonna C) nelle 6 colonne derivate:
    "Nome risorsa", "OPA@profilo", "Cliente", "Sotto progetto",
    "Riferimento tabella 1", "Commento".

    Args:
        file_excel_input: percorso del file Excel di input.
        config:           dizionario della configurazione.

    Returns:
        Tuple (df_src, df_dati_comp, col_proj, col_period,
               col_role_name, col_estimated, col_actual)
        dove df_src è il DataFrame originale e df_dati_comp è quello arricchito.
    """
    df_src = pd.read_excel(file_excel_input)

    idx_ruolo = int(config.get('ColIdxRuolo', 4))
    idx_stimato = int(config.get('ColIdxStimato', 7))
    idx_effettivo = int(config.get('ColIdxEffettivo', 8))
    idx_proj = int(config.get('ColIdxProgetto', 0))
    idx_period = int(config.get('ColIdxPeriodo', 9))

    col_role_name = df_src.columns[idx_ruolo]
    col_estimated = df_src.columns[idx_stimato]
    col_actual = df_src.columns[idx_effettivo]

    # Normalizza le ore: virgola decimale → punto, valori non numerici → 0
    for c in [col_actual, col_estimated]:
        df_src[c] = pd.to_numeric(df_src[c].astype(str).str.replace(',', '.'), errors='coerce').fillna(0.0)

    nuove_col = ["Nome risorsa", "OPA@profilo", "Cliente", "Sotto progetto",
                 "Riferimento tabella 1", "Sotto Riferimento tabella 1", "Commento"]
    df_split = pd.DataFrame(df_src.iloc[:, 2].apply(splitta_assegnazione).tolist(), columns=nuove_col)
    df_dati_comp = pd.concat([df_src, df_split], axis=1)

    col_proj = df_dati_comp.columns[idx_proj]
    col_period = df_dati_comp.columns[idx_period]

    return df_src, df_dati_comp, col_proj, col_period, col_role_name, col_estimated, col_actual

# --- CALCOLO PIVOT ---

def calcola_pivot(df_dati_comp, col_period, col_actual, col_estimated,
                  col_proj, col_role_name, weeks_limit_active, start_w, end_w):
    """Calcola le tre pivot table (actual, estimated, estimated per ruolo).

    Se weeks_limit_active è True, filtra le righe in base alla colonna
    'sett_calc' (settimana estratta dal periodo) nell'intervallo [start_w, end_w).

    Le ore vengono convertite in giornate dividendo per 8.

    Args:
        df_dati_comp:        DataFrame arricchito.
        col_period:          nome della colonna periodo (es. "CY2025-W14").
        col_actual:          nome della colonna ore consuntivate.
        col_estimated:       nome della colonna ore stimate.
        col_proj:            nome della colonna progetto.
        col_role_name:       nome della colonna ruolo.
        weeks_limit_active:  True se il filtro settimane è attivo.
        start_w:             prima settimana del filtro (inclusa).
        end_w:               ultima settimana del filtro (esclusa).

    Returns:
        Tuple (df_per_calc, pivot_actual, pivot_estimated, pivot_role_est)
        dove df_per_calc è il DataFrame filtrato usato per i calcoli.
    """
    col_sottoproj = "Sotto progetto"
    col_rif = "Riferimento tabella 1"

    df_per_calc = df_dati_comp.copy()
    if weeks_limit_active:
        mask = (df_dati_comp['sett_calc'].notna()) & \
               (df_dati_comp['sett_calc'] >= start_w) & \
               (df_dati_comp['sett_calc'] < end_w)
        df_per_calc = df_dati_comp[mask].copy()

    def create_pivot(df_input, value_col, index_cols):
        """Pivot con colonne = settimane, valori in giornate, totale riga."""
        temp_df = df_input.copy()
        temp_df[value_col] = temp_df[value_col] / 8.0
        pivot = temp_df.pivot_table(index=index_cols, columns=col_period, values=value_col, aggfunc='sum').fillna(0)
        pivot['TOTALE RIGA'] = pivot.sum(axis=1)
        return pivot.reset_index()

    index_base = [col_proj, col_sottoproj, col_rif]
    pivot_actual = create_pivot(df_per_calc, col_actual, index_base)
    pivot_estimated = create_pivot(df_per_calc, col_estimated, index_base)
    pivot_role_est = create_pivot(df_per_calc, col_estimated, [col_proj, col_sottoproj, col_role_name, col_rif])
    pivot_role_est = pivot_role_est.rename(columns={col_role_name: 'Milestone'})

    return df_per_calc, pivot_actual, pivot_estimated, pivot_role_est

# --- PREPARAZIONE RIGHE PROGETTI ---

def prepara_righe_progetti(df_src, df_dati_comp, df_per_calc, col_proj, col_actual, config, contratti_idx):
    """Costruisce la lista di dizionari da scrivere nel foglio 'progetti'.

    Per ogni progetto unico nel sorgente calcola:
    - giorni PM (ore con profilo @pm o @pc, diviso 8)
    - giorni Consulting (ore rimanenti, diviso 8)
    - giorni riscattati (da DaysRedempted<N> nel config, formato "pm, consulting")
    - opportunity, end date e riferimento da config

    Args:
        df_src:         DataFrame originale.
        df_dati_comp:   DataFrame arricchito (per leggere i metadati di riga).
        df_per_calc:    DataFrame filtrato usato per i calcoli.
        col_proj:       nome della colonna progetto.
        col_actual:     nome della colonna ore consuntivate.
        config:         dizionario della configurazione.
        contratti_idx:  indice {nome_contratto -> suffisso}.

    Returns:
        Lista di dict con chiavi A..K corrispondenti alle colonne del foglio.
    """
    rows_progetti = []
    for proj in df_src[col_proj].unique():
        df_p_full = df_dati_comp[df_dati_comp[col_proj] == proj]
        df_p_calc = df_per_calc[df_per_calc[col_proj] == proj]

        # @pm e @pc sono i profili di Project Manager / Project Coordinator
        pm_mask = df_p_calc['OPA@profilo'].str.contains('@pm|@pc', case=False, na=False)
        giorni_pm = float(df_p_calc.loc[pm_mask, col_actual].sum()) / 8.0
        giorni_cons = float(df_p_calc.loc[~pm_mask, col_actual].sum()) / 8.0

        rif_val_raw = df_p_full.iloc[0]['Riferimento tabella 1']
        rif_val = str(rif_val_raw) if pd.notna(rif_val_raw) else ""

        red_s = trova_valore_config(proj, contratti_idx, config, "DaysRedempted")
        if red_s and "," in red_s:
            try: giorni_pm_red   = float(red_s.split(",")[0].strip())
            except ValueError:   giorni_pm_red   = 0.0
            try: giorni_cons_red = float(red_s.split(",")[1].strip())
            except ValueError:   giorni_cons_red = 0.0
        else:
            giorni_pm_red, giorni_cons_red = 0.0, 0.0

        rows_progetti.append({
            'A': proj,
            'B': df_p_full.iloc[0, 1],
            'C': trova_valore_config(proj, contratti_idx, config, "Opportunity"),
            'D': trova_valore_config(proj, contratti_idx, config, "EndDate"),
            'E': giorni_pm_red,
            'F': giorni_cons_red,
            'G': giorni_pm,
            'H': giorni_cons,
            'I': "",
            'J': "",
            'K': rif_val
        })
    return rows_progetti

# --- SCRITTURA FOGLI BASE ---

def scrivi_fogli_base(file_output, df_dati_comp, rows_progetti):
    """Scrive i cinque fogli con dati grezzi nel file di output.

    I fogli Riepilogo Settimanale, Dettaglio Ruoli e Tabella di Export vengono
    creati vuoti qui; la formattazione e i dati vengono aggiunti nelle funzioni
    successive tramite openpyxl diretto.

    Il foglio 'dati' parte dalla riga 4 (righe 1-3 riservate a metadati).
    Il foglio 'progetti' parte dalla riga 11 (righe 1-10 per intestazioni).

    Args:
        file_output:    percorso del file Excel da creare/sovrascrivere.
        df_dati_comp:   DataFrame arricchito (senza la colonna temporanea sett_calc).
        rows_progetti:  lista di dict prodotta da prepara_righe_progetti.
    """
    with pd.ExcelWriter(file_output, engine='openpyxl') as writer:
        df_dati_comp.drop(columns=['sett_calc']).to_excel(writer, sheet_name='dati', index=False, startrow=0)
        pd.DataFrame(rows_progetti).to_excel(writer, sheet_name='progetti', index=False, startrow=10, header=False)
        pd.DataFrame().to_excel(writer, sheet_name='Riepilogo Settimanale', index=False)
        pd.DataFrame().to_excel(writer, sheet_name='Dettaglio Ruoli', index=False)
        pd.DataFrame().to_excel(writer, sheet_name='Tabella di Export', index=False)

# --- UTILITÀ FORMATTAZIONE ---

def autofit_columns(ws, scan_rows=12, min_width=8, max_width=60):
    """Imposta la larghezza di ogni colonna in base al testo degli header."""
    for col in ws.columns:
        max_len = min_width
        for cell in list(col)[:scan_rows]:
            if cell.value is not None:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, max_width)

# --- FORMATTAZIONE TAB PROGETTI ---

def formatta_tab_progetti(ws_p, config, rows_progetti, weeks_limit_active, bold, center):
    """Applica intestazioni, formattazione e tabella dati al foglio 'progetti'.

    La struttura del foglio è:
      - Righe 3-4  : intestazioni titolo (merge A:K)
      - Righe 5-7  : informazioni cliente (da config Intestazione8a/9a/10a)
      - Righe 9-10 : header tabella (con merge per gruppi di colonne)
      - Righe 11+  : dati progetti
      - Tabella duplicata a distanza fissa (per layout di stampa)

    Le date di scadenza entro 2 settimane vengono evidenziate in giallo.

    Args:
        ws_p:               worksheet 'progetti' openpyxl.
        config:             dizionario della configurazione.
        rows_progetti:      lista di dict prodotta da prepara_righe_progetti.
        weeks_limit_active: True se il filtro settimane è attivo.
        bold:               Font(bold=True) precostruito.
        center:             Alignment(horizontal='center') precostruito.
    """
    ws_p.merge_cells("A3:K3")
    ws_p['A3'] = config.get('Intestazione5', '')
    ws_p['A3'].font = Font(bold=True, sz=14)
    ws_p['A3'].alignment = center

    ws_p.merge_cells("A4:K4")
    ws_p['A4'] = config.get('Intestazione6', '')
    ws_p['A4'].font = Font(bold=True, sz=12)
    ws_p['A4'].alignment = center

    ws_p['A5'] = config.get('Intestazione8a', '')
    ws_p['A6'] = config.get('Intestazione9a', '')
    ws_p['A7'] = config.get('Intestazione10a', '')

    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    oggi = datetime.now().date()
    limite_due_sett = oggi + timedelta(weeks=2)

    def scrivi_intestazione(base_row):
        """Scrive l'intestazione a due righe della tabella contratti."""
        for col_lett, titolo in [('A', 'Contract name'), ('B', 'OPA Number'),
                                  ('C', 'Opportunity'), ('D', 'End Date'), ('K', 'Riferimento')]:
            ws_p[f'{col_lett}{base_row}'] = titolo
            ws_p.merge_cells(f'{col_lett}{base_row}:{col_lett}{base_row + 1}')
        ws_p.merge_cells(f'E{base_row}:F{base_row}')
        ws_p[f'E{base_row}'] = 'Days redempted'
        ws_p.merge_cells(f'G{base_row}:H{base_row}')
        ws_p[f'G{base_row}'] = 'Days Used'
        ws_p.merge_cells(f'I{base_row}:J{base_row}')
        ws_p[f'I{base_row}'] = 'Days remaining'
        for i, testo in enumerate(["Project Manager", "Consulting"] * 3):
            ws_p.cell(row=base_row + 1, column=5 + i).value = testo
        for riga in [base_row, base_row + 1]:
            for cella in ws_p[riga]:
                cella.font, cella.alignment = bold, center

    def scrivi_dati(data_start_row, arrotonda_gh=False):
        """Scrive le righe dati e applica highlight per scadenze imminenti."""
        for i, row in enumerate(rows_progetti):
            r = data_start_row + i
            ws_p.cell(row=r, column=1).value = row['A']
            ws_p.cell(row=r, column=2).value = row['B']
            ws_p.cell(row=r, column=3).value = row['C']
            ws_p.cell(row=r, column=4).value = row['D']
            ws_p.cell(row=r, column=5).value = row['E']
            ws_p.cell(row=r, column=6).value = row['F']
            g_val, h_val = row['G'], row['H']
            if arrotonda_gh:
                try: g_val = round(float(g_val))
                except (ValueError, TypeError): pass
                try: h_val = round(float(h_val))
                except (ValueError, TypeError): pass
            ws_p.cell(row=r, column=7).value = g_val
            ws_p.cell(row=r, column=8).value = h_val
            ws_p.cell(row=r, column=11).value = row['K']
            if row['D']:
                try:
                    end_date = datetime.strptime(str(row['D']), "%d/%m/%Y").date()
                    if end_date <= limite_due_sett:
                        ws_p.cell(row=r, column=4).fill = yellow_fill
                except ValueError:
                    pass
            try:
                ws_p.cell(row=r, column=9).value  = float(row['E']) - float(g_val)
            except (ValueError, TypeError):
                ws_p.cell(row=r, column=9).value  = 0.0
            try:
                ws_p.cell(row=r, column=10).value = float(row['F']) - float(h_val)
            except (ValueError, TypeError):
                ws_p.cell(row=r, column=10).value = 0.0
            ws_p.cell(row=r, column=2).alignment = center
            for col in range(6, 11):
                ws_p.cell(row=r, column=col).alignment = center
            ws_p.cell(row=r, column=11).alignment = center

    scrivi_intestazione(9)
    scrivi_dati(11, arrotonda_gh=True)

    # Tabella duplicata: distanza fissa dalla prima per il layout di stampa
    n = len(rows_progetti)
    dup_header_row = 10 + n + 6

    fill_pastello = PatternFill(fill_type="solid", fgColor="D9E1F2")
    for r in range(dup_header_row, dup_header_row + n + 2):
        for c in range(1, 12):
            ws_p.cell(row=r, column=c).fill = fill_pastello

    scrivi_intestazione(dup_header_row)
    scrivi_dati(dup_header_row + 2)

# --- FORMATTAZIONE TAB RIEPILOGO SETTIMANALE ---

def formatta_riepilogo(ws_rs, ws_dr, pivot_actual, pivot_estimated, pivot_role_est,
                       current_week_str, bold, center, green_fill, red_thick,
                       df_dati_comp, col_proj, col_role_name, col_period, col_status_k, col_status_l,
                       df_per_calc, col_actual):
    """Riempie i fogli 'Riepilogo Settimanale' e 'Dettaglio Ruoli'.

    Riepilogo Settimanale: due pivot (actual e estimated) in sequenza verticale.

    Dettaglio Ruoli: pivot estimated per ruolo con:
    - Riga 3: intervallo date di ogni settimana (lun-dom)
    - Riga 4: nome del mese (o "mese1-mese2" se la settimana è a cavallo di mese)
    - Colorazione pastello dalla settimana corrente in poi basata su
      stato schedulazione (colonne K e L del sorgente):
        verde  = Scheduled/Commit
        rosso  = Tentative/Exclude
        giallo = Tentative/Upside
        viola  = altro
    - Bordo blu spesso sulla settimana corrente
    - Colonna extra "Actual Hours (Giornate)" a destra della tabella
    - Legenda colori in fondo alla tabella

    Args:
        ws_rs:            worksheet 'Riepilogo Settimanale'.
        ws_dr:            worksheet 'Dettaglio Ruoli'.
        pivot_actual:     DataFrame pivot ore consuntivate.
        pivot_estimated:  DataFrame pivot ore stimate.
        pivot_role_est:   DataFrame pivot ore stimate per ruolo.
        current_week_str: stringa settimana corrente es. "CY2025-W14".
        bold:             Font(bold=True).
        center:           Alignment(horizontal='center').
        green_fill:       PatternFill verde chiaro per evidenziare la settimana corrente.
        red_thick:        Side(style='thick', color='FF0000') per bordi gruppi.
        df_dati_comp:     DataFrame arricchito (per lookup stato schedulazione).
        col_proj:         nome colonna progetto.
        col_role_name:    nome colonna ruolo.
        col_period:       nome colonna periodo.
        col_status_k:     nome colonna stato schedulazione (colonna K sorgente).
        col_status_l:     nome colonna stato commit/exclude (colonna L sorgente).
        df_per_calc:      DataFrame filtrato per i calcoli.
        col_actual:       nome colonna ore consuntivate.
    """
    def write_pivot(ws, start_row, df, title, border_group=False, extra_center_cols=None):
        """Scrive una pivot su un worksheet a partire da start_row.

        Aggiunge:
        - riga titolo in grassetto
        - riga header colonne in grassetto/centrato
        - righe dati con allineamento numerico centrato
        - evidenziazione verde sulla colonna della settimana corrente
        - bordi rossi spessi per raggruppare righe con stesso progetto
          (solo se border_group=True)
        - riga TOTALE SETTIMANA in fondo

        Returns:
            int: prima riga disponibile dopo la sezione (start_row + altezza + 5).
        """
        if extra_center_cols is None:
            extra_center_cols = set()
        ws.cell(row=start_row, column=1).value = title
        ws.cell(row=start_row, column=1).font = Font(bold=True, sz=12)

        for c_idx, col in enumerate(df.columns):
            cella = ws.cell(row=start_row + 1, column=c_idx + 1)
            cella.value = str(col)
            cella.font = bold
            cella.alignment = center

        num_cols_idx = len(df.columns) - len(df.select_dtypes(include=['number']).columns)
        for r_idx, row in enumerate(df.values):
            for c_idx, val in enumerate(row):
                cella = ws.cell(row=start_row + 2 + r_idx, column=c_idx + 1)
                cella.value = val
                if c_idx == 2 or c_idx >= num_cols_idx or c_idx in extra_center_cols:
                    cella.alignment = center

        ultima_riga = start_row + 2 + len(df) - 1
        ultima_col = len(df.columns)

        for r in range(start_row + 1, ultima_riga + 1):
            ws.cell(row=r, column=ultima_col).font = bold

        for c in range(1, ultima_col):
            if current_week_str in str(ws.cell(row=start_row + 1, column=c).value):
                for r in range(start_row + 1, ultima_riga + 2):
                    ws.cell(row=r, column=c).fill = green_fill

        if border_group:
            grp_start = start_row + 2
            for r in range(grp_start, ultima_riga + 1):
                if r == ultima_riga or ws.cell(row=r, column=1).value != ws.cell(row=r + 1, column=1).value:
                    for riga in range(grp_start, r + 1):
                        ws.cell(row=riga, column=1).border = Border(
                            left=red_thick,
                            top=ws.cell(row=riga, column=1).border.top,
                            bottom=ws.cell(row=riga, column=1).border.bottom
                        )
                        ws.cell(row=riga, column=ultima_col).border = Border(
                            right=red_thick,
                            top=ws.cell(row=riga, column=ultima_col).border.top,
                            bottom=ws.cell(row=riga, column=ultima_col).border.bottom
                        )
                    for col in range(1, ultima_col + 1):
                        ws.cell(row=grp_start, column=col).border = Border(
                            top=red_thick,
                            left=ws.cell(row=grp_start, column=col).border.left,
                            right=ws.cell(row=grp_start, column=col).border.right
                        )
                        ws.cell(row=r, column=col).border = Border(
                            bottom=red_thick,
                            left=ws.cell(row=r, column=col).border.left,
                            right=ws.cell(row=r, column=col).border.right
                        )
                    grp_start = r + 1

        ws.cell(row=ultima_riga + 1, column=1).value = "TOTALE SETTIMANA"
        ws.cell(row=ultima_riga + 1, column=1).font = bold
        for c in range(num_cols_idx + 1, ultima_col + 1):
            s = sum(
                ws.cell(row=r, column=c).value
                for r in range(start_row + 2, ultima_riga + 1)
                if isinstance(ws.cell(row=r, column=c).value, (int, float))
            )
            cella_tot = ws.cell(row=ultima_riga + 1, column=c)
            cella_tot.value = s
            cella_tot.font = bold
            cella_tot.alignment = center

        return ultima_riga + 5

    data_prod = datetime.now().strftime('%d/%m/%Y')
    r_next = write_pivot(ws_rs, 1, pivot_actual, f"ACTUAL HOURS (GIORNATE) - {data_prod}")
    r_next = write_pivot(ws_rs, r_next, pivot_estimated, f"ESTIMATED HOURS (GIORNATE) - {data_prod}")

    pivot_role_display = pivot_role_est.rename(columns={'Riferimento tabella 1': 'Riferimento Interno'})
    dr_next = write_pivot(ws_dr, 1, pivot_role_display,
                          f"DETTAGLIO RUOLI - ESTIMATED HOURS - {data_prod}",
                          border_group=True, extra_center_cols={1, 3})
    ultima_riga_dr = dr_next - 5

    # Inserisce riga 3 (date settimane) e riga 4 (mesi)
    ws_dr.insert_rows(3)
    ws_dr.insert_rows(4)
    num_idx = len(pivot_role_display.columns) - len(pivot_role_display.select_dtypes(include=['number']).columns)
    ultima_col = len(pivot_role_display.columns)

    # Merge colonne indice e TOTALE RIGA su righe 2-4
    for col in range(1, num_idx + 1):
        col_letter = ws_dr.cell(row=2, column=col).column_letter
        ws_dr.merge_cells(f'{col_letter}2:{col_letter}4')
        ws_dr.cell(row=2, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    last_col_letter = ws_dr.cell(row=2, column=ultima_col).column_letter
    ws_dr.merge_cells(f'{last_col_letter}2:{last_col_letter}4')
    ws_dr.cell(row=2, column=ultima_col).alignment = Alignment(horizontal='center', vertical='center')

    mesi_it = ['gennaio', 'febbraio', 'marzo', 'aprile', 'maggio', 'giugno',
               'luglio', 'agosto', 'settembre', 'ottobre', 'novembre', 'dicembre']

    current_week_col_dr = None
    for col in range(num_idx + 1, ultima_col):
        header_val = ws_dr.cell(row=2, column=col).value
        if not header_val:
            continue
        if current_week_str in str(header_val) and current_week_col_dr is None:
            current_week_col_dr = col
        # Riga 3: date settimana
        date_str = calcola_date_settimana(str(header_val))
        if date_str:
            c3 = ws_dr.cell(row=3, column=col)
            c3.value = date_str
            c3.alignment = center
            c3.font = bold
        # Riga 4: mesi
        match = re.search(r'(\d{4})-W(\d+)', str(header_val), re.IGNORECASE)
        if match:
            yr, wk = int(match.group(1)), int(match.group(2))
            try:
                lun = datetime.fromisocalendar(yr, wk, 1)
                dom = datetime.fromisocalendar(yr, wk, 7)
                ml, md = mesi_it[lun.month - 1], mesi_it[dom.month - 1]
                c4 = ws_dr.cell(row=4, column=col)
                c4.value = ml if ml == md else f"{ml}-{md}"
                c4.alignment = center
                c4.font = bold
            except Exception:
                pass

    fill_verde_past  = PatternFill(fill_type="solid", fgColor="C8E6C9")
    fill_rosso_past  = PatternFill(fill_type="solid", fgColor="FFCDD2")
    fill_giallo_past = PatternFill(fill_type="solid", fgColor="FFF9C4")
    fill_viola_past  = PatternFill(fill_type="solid", fgColor="E1BEE7")
    no_fill          = PatternFill(fill_type=None)

    data_start_row = 5
    ultima_riga_ws = ultima_riga_dr + 2
    col_rif_src    = 'Riferimento tabella 1'

    # Rimuove green_fill dalla colonna settimana corrente e aggiunge bordi blu spessi
    if current_week_col_dr is not None:
        blue_thick = Side(style='thick', color='0070C0')
        first_week_col = num_idx + 1
        for r in range(2, ultima_riga_ws + 2):
            # Bordo sinistro e destro sulla settimana corrente
            cw = ws_dr.cell(row=r, column=current_week_col_dr)
            cw.fill = no_fill
            eb = cw.border
            cw.border = Border(top=eb.top, bottom=eb.bottom, left=blue_thick, right=blue_thick)
            # Bordo destro sulla colonna prima della settimana corrente
            if current_week_col_dr > 1:
                cp = ws_dr.cell(row=r, column=current_week_col_dr - 1)
                ep = cp.border
                cp.border = Border(top=ep.top, bottom=ep.bottom, left=ep.left, right=blue_thick)
            # Bordo sinistro sulla colonna dopo la settimana corrente
            cn = ws_dr.cell(row=r, column=current_week_col_dr + 1)
            en = cn.border
            cn.border = Border(top=en.top, bottom=en.bottom, left=blue_thick, right=en.right)
            # Bordo prima della prima colonna settimana (se diversa dalla corrente)
            if first_week_col != current_week_col_dr:
                cf = ws_dr.cell(row=r, column=first_week_col)
                ef = cf.border
                cf.border = Border(top=ef.top, bottom=ef.bottom, left=blue_thick, right=ef.right)
                ci = ws_dr.cell(row=r, column=first_week_col - 1)
                ei = ci.border
                ci.border = Border(top=ei.top, bottom=ei.bottom, left=ei.left, right=blue_thick)

    # Colorazione pastello dalla settimana corrente in poi (inclusa)
    if current_week_col_dr is not None:
        for c in range(current_week_col_dr, ultima_col):
            week_val = str(ws_dr.cell(row=2, column=c).value or '')
            if not week_val:
                continue
            for r in range(data_start_row, ultima_riga_ws + 1):
                proj      = ws_dr.cell(row=r, column=1).value
                if not proj:
                    continue
                sottoprog = ws_dr.cell(row=r, column=2).value
                milestone = ws_dr.cell(row=r, column=3).value
                rif       = ws_dr.cell(row=r, column=4).value

                mask = (
                    (df_dati_comp[col_proj].astype(str)        == str(proj))      &
                    (df_dati_comp['Sotto progetto'].astype(str) == str(sottoprog)) &
                    (df_dati_comp[col_role_name].astype(str)   == str(milestone)) &
                    (df_dati_comp[col_rif_src].astype(str)     == str(rif))       &
                    (df_dati_comp[col_period].astype(str)      == week_val)
                )
                subset = df_dati_comp[mask]
                if subset.empty:
                    continue

                k_vals = set(subset[col_status_k].astype(str).str.strip().str.lower().unique())
                l_vals = set(subset[col_status_l].astype(str).str.strip().str.lower().unique())

                cella = ws_dr.cell(row=r, column=c)
                if k_vals == {'scheduled'} and l_vals == {'commit'}:
                    cella.fill = fill_verde_past
                elif k_vals == {'tentative'} and l_vals == {'exclude'}:
                    cella.fill = fill_rosso_past
                elif k_vals == {'tentative'} and l_vals == {'upside'}:
                    cella.fill = fill_giallo_past
                else:
                    cella.fill = fill_viola_past

    # Colonna Actual Hours (Giornate) — ultima_col + 2 (lascia una colonna vuota)
    act_col = ultima_col + 2
    actual_grp = (
        df_per_calc.groupby([col_proj, 'Sotto progetto', col_role_name, col_rif_src])[col_actual]
        .sum() / 8.0
    ).to_dict()

    act_hdr = ws_dr.cell(row=2, column=act_col)
    act_hdr.value = "Actual Hours (Giornate)"
    act_hdr.font = bold
    act_hdr.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_dr.merge_cells(
        start_row=2, start_column=act_col,
        end_row=4,   end_column=act_col
    )

    totale_actual = 0.0
    for r in range(data_start_row, ultima_riga_ws + 1):
        proj      = ws_dr.cell(row=r, column=1).value
        if not proj:
            continue
        sottoprog = ws_dr.cell(row=r, column=2).value
        milestone = ws_dr.cell(row=r, column=3).value
        rif       = ws_dr.cell(row=r, column=4).value
        val = actual_grp.get((proj, sottoprog, milestone, rif), 0.0)
        c_act = ws_dr.cell(row=r, column=act_col)
        c_act.value = val
        c_act.alignment = center
        totale_actual += val

    tot_act = ws_dr.cell(row=ultima_riga_ws + 1, column=act_col)
    tot_act.value = totale_actual
    tot_act.font = bold
    tot_act.alignment = center

    # Legenda colori sotto la tabella a partire dalla colonna settimana corrente
    if current_week_col_dr is not None:
        green_med  = Side(style='medium', color='00B050')
        green_thin = Side(style='thin',   color='00B050')
        leggenda = [
            ("Scheduled", "Commit",  fill_verde_past),
            ("Tentative",  "Exclude", fill_rosso_past),
            ("Tentative",  "Upside",  fill_giallo_past),
            ("Altro",      "",        fill_viola_past),
        ]
        leg_start = ultima_riga_ws + 3
        lc1, lc2  = current_week_col_dr, current_week_col_dr + 1
        n_leg     = len(leggenda)

        for i, (t1, t2, fill) in enumerate(leggenda):
            r        = leg_start + i
            is_top   = (i == 0)
            is_bot   = (i == n_leg - 1)
            for off, txt in [(0, t1), (1, t2)]:
                c   = lc1 + off
                cel = ws_dr.cell(row=r, column=c)
                cel.value     = txt
                cel.fill      = fill
                cel.alignment = center
                cel.font      = bold
                cel.border    = Border(
                    top    = green_med  if is_top else green_thin,
                    bottom = green_med  if is_bot else green_thin,
                    left   = green_med  if off == 0 else green_thin,
                    right  = green_med  if off == 1 else green_thin,
                )
        # Merge cella "Altro" sulle due colonne
        ws_dr.merge_cells(
            start_row=leg_start + n_leg - 1, start_column=lc1,
            end_row=leg_start + n_leg - 1,   end_column=lc2
        )

# --- FORMATTAZIONE TAB EXPORT ---

def formatta_tab_export(ws_e, config, df_per_calc, col_rif, col_actual,
                        fill_verde, fill_nero, font_bianco_bold, center, right_align):
    """Riempie il foglio 'Tabella di Export' con il riepilogo giornate per codice ordine.

    Legge le righe di esportazione dalla config (Export3, Export4, …): ogni riga
    è una lista di valori separati da virgola che rappresentano i campi della
    tabella (codice interno, codice offerta, descrizione, …).

    Per ogni riga cerca se uno dei valori del campo "codice" è presente nella
    colonna "Riferimento tabella 1" del sorgente e, se sì, riporta la somma
    delle ore consuntivate (in giornate) nella colonna J.

    Come per il foglio progetti, la tabella è duplicata: la prima copia ha
    valori esatti, la seconda arrotonda le giornate all'intero più vicino.

    Args:
        ws_e:             worksheet 'Tabella di Export'.
        config:           dizionario della configurazione.
        df_per_calc:      DataFrame filtrato per i calcoli.
        col_rif:          nome della colonna riferimento (es. "Riferimento tabella 1").
        col_actual:       nome della colonna ore consuntivate.
        fill_verde:       PatternFill verde scuro per la riga titolo.
        fill_nero:        PatternFill nero per le righe header.
        font_bianco_bold: Font bianco grassetto per testo su sfondo scuro.
        center:           Alignment(horizontal='center').
        right_align:      Alignment(horizontal='right').

    Returns:
        int: numero di riga dell'ultima riga scritta nella tabella duplicata.
    """
    somme_rif = (df_per_calc.groupby(col_rif)[col_actual].sum() / 8.0).to_dict()
    somme_rif = {str(k).strip(): v for k, v in somme_rif.items()}

    col_sotto_rif = "Sotto Riferimento tabella 1"
    somme_sotto_rif = (df_per_calc.groupby(col_sotto_rif)[col_actual].sum() / 8.0).to_dict()
    somme_sotto_rif = {str(k).strip(): v for k, v in somme_sotto_rif.items()}
    log.info("Riferimenti disponibili: %s | Sotto-rif: %s",
             list(somme_rif.keys()), list(somme_sotto_rif.keys()))

    righe_export = []
    idx = 3
    while f"Export{idx}" in config:
        vals = [v.strip() for v in config[f"Export{idx}"].split(',')]
        ref_key = vals[-1] if vals else ''
        valore_match = (
            somme_rif.get(ref_key, 0.0) + somme_sotto_rif.get(ref_key, 0.0)
        ) if ref_key else 0.0
        log.info("Export%d: ref=%r → %.2f gg (rif=%.2f + sotto=%.2f)",
                 idx, ref_key, valore_match,
                 somme_rif.get(ref_key, 0.0), somme_sotto_rif.get(ref_key, 0.0))
        righe_export.append((vals, valore_match))
        idx += 1

    def scrivi_titolo(start_row):
        ws_e.merge_cells(f"B{start_row}:M{start_row}")
        ws_e[f'B{start_row}'] = f"{config.get('Export1', '')} {datetime.now().strftime('%d/%m/%Y')}"
        ws_e[f'B{start_row}'].fill = fill_verde
        ws_e[f'B{start_row}'].font = font_bianco_bold
        ws_e[f'B{start_row}'].alignment = center

    def scrivi_header(start_row):
        header_vals = config.get('Export2', '').split(',')
        for i in range(1, 14):
            cella = ws_e.cell(row=start_row, column=i)
            if i <= len(header_vals):
                cella.value = header_vals[i - 1].strip()
            cella.fill = fill_nero
            cella.font = font_bianco_bold
            cella.alignment = center

    def scrivi_dati(start_row, arrotonda=False):
        for i, (vals, valore_match) in enumerate(righe_export):
            r = start_row + i
            for j, v in enumerate(vals):
                ws_e.cell(row=r, column=1 + j).value = v
            if arrotonda:
                try:
                    i_raw = ws_e.cell(row=r, column=9).value
                    ws_e.cell(row=r, column=9).value = round(float(str(i_raw).replace(',', '.')))
                except (ValueError, TypeError):
                    pass
            ws_e.cell(row=r, column=9).alignment = right_align
            j_val = round(valore_match) if arrotonda else valore_match
            ws_e.cell(row=r, column=10).value = j_val
            try:
                i_num = float(str(ws_e.cell(row=r, column=9).value or 0).replace(',', '.'))
                k_val = round(i_num - j_val) if arrotonda else round(i_num - j_val, 2)
                ws_e.cell(row=r, column=11).value = k_val
            except (ValueError, TypeError):
                ws_e.cell(row=r, column=11).value = ''
        return start_row + len(righe_export)

    fill_pastello = PatternFill(fill_type="solid", fgColor="D9E1F2")
    n_righe = len(righe_export)
    dup_start_pre = 3 + n_righe + 5

    # Prima tabella (decimale, righe 1..dup_start_pre-1): tutte le colonne A-N
    for r in range(1, dup_start_pre):
        for c in range(1, 15):
            ws_e.cell(row=r, column=c).fill = fill_pastello

    # Seconda tabella (arrotondata): solo colonne A e N
    for r in range(dup_start_pre, dup_start_pre + 2 + n_righe):
        ws_e.cell(row=r, column=1).fill = fill_pastello
        ws_e.cell(row=r, column=14).fill = fill_pastello

    scrivi_titolo(1)
    scrivi_header(2)
    riga_fine = scrivi_dati(3)

    dup_start = riga_fine + 5
    scrivi_titolo(dup_start)
    scrivi_header(dup_start + 1)
    scrivi_dati(dup_start + 2, arrotonda=True)

    return dup_start + 1 + len(righe_export)

# --- AGGIUNGI NOTE ---

def aggiungi_note(ws_p, ws_e, anno_corrente, start_w, end_w, rows_progetti, riga_export, bold):
    """Aggiunge una nota sul filtro settimane attivo in fondo ai fogli 'progetti' ed 'Export'.

    La nota riporta l'intervallo di settimane e le date corrispondenti.

    Args:
        ws_p:           worksheet 'progetti'.
        ws_e:           worksheet 'Tabella di Export'.
        anno_corrente:  anno usato per calcolare le date.
        start_w:        prima settimana del filtro.
        end_w:          ultima settimana del filtro.
        rows_progetti:  lista righe progetto (per calcolare la posizione).
        riga_export:    ultima riga scritta nel foglio Export.
        bold:           Font(bold=True).
    """
    d_s, d_e = get_date_range(anno_corrente, start_w, end_w)
    nota1 = "NOTA: Limiti settimane ATTIVATI:"
    nota2 = f"    Range: {start_w} - {end_w} [{d_s} - {d_e}]"
    n = len(rows_progetti)
    lr_p = 17 + 2 * n  # dopo la tabella duplicata (10+n+6+2+n-1)
    for ws_t, lr_t in [(ws_p, lr_p), (ws_e, riga_export)]:
        ws_t[f'A{lr_t + 2}'] = nota1
        ws_t[f'A{lr_t + 2}'].font = Font(italic=True)
        ws_t[f'A{lr_t + 3}'] = nota2
        ws_t[f'A{lr_t + 3}'].font = bold

# --- GRAFICI EXCEL ---

def _vline(lc, ws, col, data_row, n_weeks, color_hex, err_val=100.0):
    """Simula una linea verticale nel line chart tramite error bar Y su serie nascosta."""
    ws.cell(row=1, column=col).value = ''
    for r in range(2, n_weeks + 2):
        ws.cell(row=r, column=col).value = None
    ws.cell(row=data_row, column=col).value = 0.0
    lc.add_data(Reference(ws, min_col=col, min_row=2, max_row=1 + n_weeks))
    ser = lc.series[-1]
    try:
        from openpyxl.chart.marker import Marker
        ser.marker = Marker(symbol='none')
    except Exception:
        pass
    try:
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.line import LineProperties
        gp = GraphicalProperties()
        gp.ln = LineProperties(noFill=True)
        ser.spPr = gp
    except Exception:
        pass
    try:
        from openpyxl.chart.error_bar import ErrorBars
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.line import LineProperties
        gp2 = GraphicalProperties()
        lp2 = LineProperties()
        lp2.solidFill = color_hex
        gp2.ln = lp2
        ser.errBars = ErrorBars(
            errDir='y', errBarType='both', errValType='fixedVal',
            val=err_val, noEndCap=True, spPr=gp2
        )
    except Exception:
        pass


def _week_vline_indices(all_weeks):
    """Ritorna (cw_idx, me_idx_list): indice settimana corrente e lista indici fine mese."""
    from datetime import datetime as _dt
    _oggi = _dt.now()
    _cw_label = f"CY{_oggi.year}-W{_oggi.isocalendar()[1]:02d}"
    _weeks_str = [str(w) for w in all_weeks]
    try:
        _cw_idx = _weeks_str.index(_cw_label)
    except ValueError:
        _cw_idx = -1
    _me_idx = []
    for _i in range(len(_weeks_str) - 1):
        _m1 = re.search(r'(\d{4}).*W(\d+)', _weeks_str[_i], re.IGNORECASE)
        _m2 = re.search(r'(\d{4}).*W(\d+)', _weeks_str[_i + 1], re.IGNORECASE)
        if _m1 and _m2:
            try:
                from datetime import datetime as _dt2
                _d1 = _dt2.fromisocalendar(int(_m1.group(1)), int(_m1.group(2)), 1)
                _d2 = _dt2.fromisocalendar(int(_m2.group(1)), int(_m2.group(2)), 1)
                if (_d1.year, _d1.month) != (_d2.year, _d2.month):
                    _me_idx.append(_i)
            except Exception:
                pass
    return _cw_idx, _me_idx


def crea_grafici_rh(wb, rows_progetti, df_per_calc, col_proj, col_period, col_actual, bold, center):
    """Crea il foglio 'Grafici RH': line chart giornate rimaste per contratto nel tempo."""
    ws = wb.create_sheet("Grafici RH")

    def _sk(s):
        m = re.search(r'(\d{4}).*W(\d+)', str(s), re.IGNORECASE)
        return (int(m.group(1)), int(m.group(2))) if m else (0, 0)

    all_weeks = sorted(df_per_calc[col_period].dropna().unique(), key=_sk)
    n_weeks, n_proj = len(all_weeks), len(rows_progetti)
    if not all_weeks or not rows_progetti:
        return

    # Pivot project × week → giorni consumati cumulati
    pivot = (
        df_per_calc.groupby([col_proj, col_period])[col_actual]
        .sum().unstack(fill_value=0)
        .reindex(columns=all_weeks, fill_value=0)
        / 8.0
    )
    cumsum = pivot.cumsum(axis=1)

    # Tabella dati: col 1 = settimana, col 2+ = rimasti per progetto
    ws.cell(row=1, column=1).value = "Settimana"
    ws.cell(row=1, column=1).font = bold
    ws.cell(row=1, column=1).alignment = center
    for j, row in enumerate(rows_progetti, 2):
        c = ws.cell(row=1, column=j)
        c.value = row['A']
        c.font = bold
        c.alignment = center

    for i, week in enumerate(all_weeks, 2):
        ws.cell(row=i, column=1).value = str(week)
        for j, row in enumerate(rows_progetti, 2):
            proj = row['A']
            total_red = float(row['E'] or 0) + float(row['F'] or 0)
            cum = float(cumsum.loc[proj, week]) if proj in cumsum.index else 0.0
            ws.cell(row=i, column=j).value = round(total_red - cum, 2)

    # Calcola il massimo delle giornate rimaste per impostare i bounds dell'asse Y
    axis_max = 10.0
    for row in rows_progetti:
        proj = row['A']
        total_red = float(row['E'] or 0) + float(row['F'] or 0)
        if proj in cumsum.index:
            axis_max = max(axis_max, float((total_red - cumsum.loc[proj]).max()))
        else:
            axis_max = max(axis_max, total_red)
    axis_max = axis_max * 1.1

    # Line chart: ascisse = settimane, ordinate = giornate rimaste, una linea per contratto
    lc = LineChart()
    lc.title = "Giorni Rimasti per Contratto nel Tempo"
    lc.y_axis.title = "Giornate Rimaste"
    lc.x_axis.title = "Settimana"
    lc.width = 30
    lc.height = 16
    try:
        lc.y_axis.scaling.min = 0
        lc.y_axis.scaling.max = axis_max
    except Exception:
        pass

    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + n_weeks)
    for j in range(2, 2 + n_proj):
        ref = Reference(ws, min_col=j, min_row=1, max_row=1 + n_weeks)
        lc.add_data(ref, titles_from_data=True)
    lc.set_categories(cats)

    _cw_idx, _me_idx = _week_vline_indices(all_weeks)
    _col_vl = 2 + n_proj
    if 0 <= _cw_idx < n_weeks:
        _vline(lc, ws, _col_vl, _cw_idx + 2, n_weeks, 'DC2626', axis_max * 100)
        _col_vl += 1
    for _mi in _me_idx:
        if 0 <= _mi < n_weeks:
            _vline(lc, ws, _col_vl, _mi + 2, n_weeks, '000000', axis_max * 100)
            _col_vl += 1

    ws.add_chart(lc, f"A{n_weeks + 4}")

    autofit_columns(ws)


def crea_grafici_intesa(wb, df_per_calc, col_period, col_actual, col_rif, config, bold, center):
    """Crea il foglio 'Grafici Intesa': line chart giornate rimaste per voce nel tempo."""
    ws = wb.create_sheet("Grafici Intesa")

    def _sk(s):
        m = re.search(r'(\d{4}).*W(\d+)', str(s), re.IGNORECASE)
        return (int(m.group(1)), int(m.group(2))) if m else (0, 0)

    all_weeks = sorted(df_per_calc[col_period].dropna().unique(), key=_sk)
    n_weeks = len(all_weeks)
    if not all_weeks:
        return

    col_sotto_rif = "Sotto Riferimento tabella 1"

    # Voci dalla config (skip separatori '-')
    voci = []
    idx = 3
    while f"Export{idx}" in config:
        vals = [v.strip() for v in config[f"Export{idx}"].split(',')]
        if vals and vals[0] != '-':
            desc = vals[2] if len(vals) > 2 and vals[2] else vals[0]
            ref_key = str(vals[-1]) if vals else ''
            try:
                acq = float(vals[8] if len(vals) > 8 else '0')
            except (ValueError, TypeError):
                acq = 0.0
            voci.append((desc, ref_key, acq))
        idx += 1

    if not voci:
        return

    n_voci = len(voci)

    # Cumsum settimanale per ogni ref_key (rif + sotto-rif)
    def _weekly(ref_key):
        mask = (
            (df_per_calc[col_rif].astype(str).str.strip() == ref_key) |
            (df_per_calc[col_sotto_rif].astype(str).str.strip() == ref_key)
        )
        return df_per_calc[mask].groupby(col_period)[col_actual].sum() / 8.0

    cumsum_voci = {}
    for _, ref_key, _ in voci:
        wc = _weekly(ref_key)
        cum, cs = 0.0, {}
        for w in all_weeks:
            cum += float(wc.get(w, 0.0))
            cs[w] = cum
        cumsum_voci[ref_key] = cs

    # Tabella dati: col 1 = settimana, col 2+ = rimasti per voce
    ws.cell(row=1, column=1).value = "Settimana"
    ws.cell(row=1, column=1).font = bold
    ws.cell(row=1, column=1).alignment = center
    for j, (desc, _, _) in enumerate(voci, 2):
        c = ws.cell(row=1, column=j)
        c.value = desc
        c.font = bold
        c.alignment = center

    for i, week in enumerate(all_weeks, 2):
        ws.cell(row=i, column=1).value = str(week)
        for j, (_, ref_key, acq) in enumerate(voci, 2):
            ws.cell(row=i, column=j).value = round(acq - cumsum_voci[ref_key].get(week, 0.0), 2)

    # Calcola il massimo delle giornate rimaste per impostare i bounds dell'asse Y
    axis_max = 10.0
    for _, ref_key, acq in voci:
        for w in all_weeks:
            axis_max = max(axis_max, acq - cumsum_voci[ref_key].get(w, 0.0))
    axis_max = axis_max * 1.1

    # Line chart: ascisse = settimane, ordinate = giornate rimaste, una linea per voce
    lc = LineChart()
    lc.title = "Giorni Rimasti per Voce nel Tempo"
    lc.y_axis.title = "Giornate Rimaste"
    lc.x_axis.title = "Settimana"
    lc.width = 30
    lc.height = 16
    try:
        lc.y_axis.scaling.min = 0
        lc.y_axis.scaling.max = axis_max
    except Exception:
        pass

    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + n_weeks)
    for j in range(2, 2 + n_voci):
        ref = Reference(ws, min_col=j, min_row=1, max_row=1 + n_weeks)
        lc.add_data(ref, titles_from_data=True)
    lc.set_categories(cats)

    _cw_idx, _me_idx = _week_vline_indices(all_weeks)
    _col_vl = 2 + n_voci
    if 0 <= _cw_idx < n_weeks:
        _vline(lc, ws, _col_vl, _cw_idx + 2, n_weeks, 'DC2626', axis_max * 100)
        _col_vl += 1
    for _mi in _me_idx:
        if 0 <= _mi < n_weeks:
            _vline(lc, ws, _col_vl, _mi + 2, n_weeks, '000000', axis_max * 100)
            _col_vl += 1

    ws.add_chart(lc, f"A{n_weeks + 4}")

    autofit_columns(ws)


# --- GENERAZIONE HTML ---

def genera_html(df_dati_comp, rows_progetti, pivot_actual, pivot_estimated,
                pivot_role_est, df_per_calc, config, col_rif, col_actual,
                col_proj, col_period, col_estimated, file_output):
    """Genera un file HTML navigabile con tabelle e grafici Chart.js.

    Il file ha lo stesso nome del file Excel con estensione .html e viene
    scritto nella stessa directory. Include una barra laterale fissa con
    indice e link ancorati a ciascuna sezione, più 4 grafici interattivi.
    """
    html_path = os.path.splitext(file_output)[0] + '.html'
    now_str   = datetime.now().strftime('%d/%m/%Y %H:%M')
    titolo    = config.get('Intestazione5', 'Report Settimanale Risorse')

    df_dati = df_dati_comp.drop(columns=['sett_calc'], errors='ignore')

    def _f(v):
        try:    return round(float(v), 2)
        except: return v or ''

    # ── DataFrame progetti ───────────────────────────────────────────────────
    df_proj = pd.DataFrame([{
        'Contract Name': r['A'],
        'OPA Number':    r['B'],
        'Opportunity':   r['C'],
        'End Date':      r['D'],
        'Risc. PM':      _f(r['E']),
        'Risc. Cons.':   _f(r['F']),
        'Usati PM':      _f(r['G']),
        'Usati Cons.':   _f(r['H']),
        'Rem. PM':       _f(r['E'] - r['G']),
        'Rem. Cons.':    _f(r['F'] - r['H']),
        'Riferimento':   r['K'],
    } for r in rows_progetti])

    # ── DataFrame export ─────────────────────────────────────────────────────
    somme_rif = {str(k).strip(): round(v / 8.0, 2)
                 for k, v in df_per_calc.groupby(col_rif)[col_actual].sum().items()}
    somme_sotto_rif_html = {str(k).strip(): round(v / 8.0, 2)
                            for k, v in df_per_calc.groupby("Sotto Riferimento tabella 1")[col_actual].sum().items()}
    hdr_exp = [h.strip() for h in config.get('Export2', '').split(',')]
    righe_exp = []
    i_e = 3
    while f"Export{i_e}" in config:
        vals = [v.strip() for v in config[f"Export{i_e}"].split(',')]
        ref_key = vals[-1] if vals else ''
        gg = (somme_rif.get(ref_key, 0.0) + somme_sotto_rif_html.get(ref_key, 0.0)) if ref_key else 0.0
        try:
            i_num = float(str(vals[8] if len(vals) > 8 else '0').replace(',', '.'))
            k_val = round(i_num - gg, 2)
        except (ValueError, TypeError):
            k_val = ''
        while len(vals) < 9:
            vals.append('')
        righe_exp.append(vals[:9] + [round(gg, 2)] + [k_val])
        i_e += 1

    if righe_exp:
        nc     = max(len(r) for r in righe_exp)
        df_exp = pd.DataFrame(
            [r + [''] * (nc - len(r)) for r in righe_exp],
            columns=(hdr_exp + [''] * nc)[:nc]
        )
    else:
        df_exp = pd.DataFrame()

    pv_ruoli = pivot_role_est.rename(columns={'Riferimento tabella 1': 'Riferimento'})

    # ── Dati grafici ─────────────────────────────────────────────────────────
    actual_pp    = df_per_calc.groupby(col_proj)[col_actual].sum() / 8.0
    estimated_pp = df_per_calc.groupby(col_proj)[col_estimated].sum() / 8.0
    progetti_list = list(actual_pp.index)

    def _sort_key(s):
        m = re.search(r'(\d{4}).*W(\d+)', str(s), re.IGNORECASE)
        return (int(m.group(1)), int(m.group(2))) if m else (0, 0)

    trend = df_per_calc.groupby(col_period)[col_actual].sum() / 8.0
    trend = trend.loc[sorted(trend.index, key=_sort_key)]

    risorse_s = (
        df_per_calc[df_per_calc['Nome risorsa'].astype(str).str.strip().ne('') &
                    df_per_calc['Nome risorsa'].notna()]
        .groupby('Nome risorsa')[col_actual].sum() / 8.0
    ).sort_values(ascending=False).head(10)

    # ── Time-series giornate rimaste per Grafici RH e Intesa ─────────────────
    all_weeks_ts = sorted(df_per_calc[col_period].dropna().unique(), key=_sort_key)

    rh_pivot_ts = (
        df_per_calc.groupby([col_proj, col_period])[col_actual]
        .sum().unstack(fill_value=0)
        .reindex(columns=all_weeks_ts, fill_value=0)
        / 8.0
    ).cumsum(axis=1)

    rh_series = []
    for row in rows_progetti:
        proj = row['A']
        total_red = float(row['E'] or 0) + float(row['F'] or 0)
        rh_series.append({'label': proj, 'data': [
            round(total_red - (float(rh_pivot_ts.loc[proj, w]) if proj in rh_pivot_ts.index else 0.0), 2)
            for w in all_weeks_ts
        ]})

    col_sotto_rif_ts = "Sotto Riferimento tabella 1"
    intesa_series = []
    _ie = 3
    while f"Export{_ie}" in config:
        _v = [v.strip() for v in config[f"Export{_ie}"].split(',')]
        _ie += 1
        if not _v or _v[0] == '-':
            continue
        _desc = _v[2] if len(_v) > 2 and _v[2] else _v[0]
        _rk   = str(_v[-1]) if _v else ''
        try: _acq = float(_v[8] if len(_v) > 8 else '0')
        except (ValueError, TypeError): _acq = 0.0
        _wc = df_per_calc[
            (df_per_calc[col_rif].astype(str).str.strip() == _rk) |
            (df_per_calc[col_sotto_rif_ts].astype(str).str.strip() == _rk)
        ].groupby(col_period)[col_actual].sum() / 8.0
        _cum, _data = 0.0, []
        for w in all_weeks_ts:
            _cum += float(_wc.get(w, 0.0))
            _data.append(round(_acq - _cum, 2))
        intesa_series.append({'label': _desc, 'data': _data})

    # ── Meta: settimana corrente e fine mesi per annotazioni ─────────────────
    _oggi = datetime.now()
    _cw_label = f"CY{_oggi.year}-W{_oggi.isocalendar()[1]:02d}"
    _weeks_str = [str(w) for w in all_weeks_ts]
    try: _cw_idx = _weeks_str.index(_cw_label)
    except ValueError: _cw_idx = -1
    _me_pos = []
    for _i in range(len(all_weeks_ts) - 1):
        _m1 = re.search(r'(\d{4}).*W(\d+)', _weeks_str[_i],     re.IGNORECASE)
        _m2 = re.search(r'(\d{4}).*W(\d+)', _weeks_str[_i + 1], re.IGNORECASE)
        if _m1 and _m2:
            try:
                _d1 = datetime.fromisocalendar(int(_m1.group(1)), int(_m1.group(2)), 1)
                _d2 = datetime.fromisocalendar(int(_m2.group(1)), int(_m2.group(2)), 1)
                if (_d1.year, _d1.month) != (_d2.year, _d2.month):
                    _me_pos.append(_i + 0.5)
            except Exception:
                pass

    chart_data = {
        'ae': {
            'labels':    progetti_list,
            'actual':    [round(float(actual_pp[p]), 2) for p in progetti_list],
            'estimated': [round(float(estimated_pp.get(p, 0.0)), 2) for p in progetti_list],
        },
        'trend': {
            'labels': [str(w) for w in trend.index],
            'values': [round(float(v), 2) for v in trend.values],
        },
        'risorse': {
            'labels': list(risorse_s.index),
            'values': [round(float(v), 2) for v in risorse_s.values],
        },
        'contratti': {
            'labels':     [r['A'] for r in rows_progetti],
            'risc_pm':    [float(r['E'] or 0) for r in rows_progetti],
            'risc_cons':  [float(r['F'] or 0) for r in rows_progetti],
            'usati_pm':   [float(r['G'] or 0) for r in rows_progetti],
            'usati_cons': [float(r['H'] or 0) for r in rows_progetti],
        },
        'rh': {
            'weeks':  [str(w) for w in all_weeks_ts],
            'series': rh_series,
        },
        'intesa': {
            'weeks':  [str(w) for w in all_weeks_ts],
            'series': intesa_series,
        },
        'meta': {
            'current_week_idx':    _cw_idx,
            'month_end_positions': _me_pos,
        },
    }
    chart_json = json.dumps(chart_data, ensure_ascii=False)

    # ── CSS ─────────────────────────────────────────────────────────────────
    css = """
:root{--pri:#1a56db;--pri-l:#dbeafe;--pri-d:#1e40af;--sid:#0f172a;
  --bg:#f1f5f9;--srf:#fff;--brd:#e2e8f0;--txt:#1e293b;--mut:#64748b}
*,::before,::after{box-sizing:border-box;margin:0;padding:0}
html{scroll-behavior:smooth}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
  background:var(--bg);color:var(--txt);display:flex;min-height:100vh;font-size:14px}
.sidebar{width:215px;min-width:215px;background:var(--sid);padding:1.25rem 1rem;
  position:sticky;top:0;height:100vh;overflow-y:auto;
  display:flex;flex-direction:column;gap:1.25rem}
.sb-title{color:#fff;font-size:.9rem;font-weight:700;padding-bottom:.875rem;
  border-bottom:1px solid rgba(255,255,255,.1)}
.sb-meta{color:#475569;font-size:.73rem;line-height:1.65}
.nav-list{list-style:none;display:flex;flex-direction:column;gap:1px}
.nav-sep{color:#374151;font-size:.65rem;font-weight:700;letter-spacing:.09em;
  text-transform:uppercase;padding:.75rem .5rem .2rem;margin-top:.25rem}
.nav-list a{color:#94a3b8;text-decoration:none;display:block;padding:.37rem .65rem;
  border-radius:4px;font-size:.83rem;transition:background .15s,color .15s}
.nav-list a:hover{background:rgba(255,255,255,.08);color:#e2e8f0}
.nav-list a.sub{padding-left:1.15rem;font-size:.79rem;color:#64748b}
.nav-list a.sub:hover{color:#94a3b8}
main{flex:1;padding:1.75rem 2rem;overflow-x:hidden;min-width:0}
.pg-hdr{margin-bottom:1.75rem}
.pg-hdr h1{font-size:1.45rem;font-weight:700;color:var(--pri-d);margin-bottom:.3rem}
.pg-hdr p{color:var(--mut);font-size:.82rem}
.sec{background:var(--srf);border-radius:8px;
  box-shadow:0 1px 4px rgba(0,0,0,.09);margin-bottom:1.75rem;overflow:hidden}
.sec-hdr{background:var(--pri);color:#fff;padding:.7rem 1.2rem;
  display:flex;align-items:center;justify-content:space-between}
.sec-hdr h2{font-size:.95rem;font-weight:600;letter-spacing:.01em}
.badge{background:rgba(255,255,255,.22);border-radius:20px;
  padding:.1rem .55rem;font-size:.72rem}
.sub-sec{padding:.875rem 1.2rem;border-bottom:1px solid var(--brd)}
.sub-sec:last-child{border-bottom:none}
.sub-sec h3{font-size:.85rem;font-weight:600;color:var(--pri-d);margin-bottom:.65rem;
  display:flex;align-items:center;gap:.45rem}
.sub-sec h3::before{content:'';display:inline-block;width:3px;height:.9em;
  background:var(--pri);border-radius:2px}
.tbl-wrap{overflow:auto;max-height:520px;border-top:1px solid var(--brd)}
table{border-collapse:collapse;width:100%;font-size:.8rem}
thead th{background:var(--pri-l);color:var(--pri-d);font-weight:600;padding:.5rem .72rem;
  text-align:left;white-space:nowrap;position:sticky;top:0;
  border-bottom:2px solid var(--pri);z-index:1}
tbody tr:nth-child(even){background:#f8fafc}
tbody tr:hover{background:#eff6ff}
td{padding:.38rem .72rem;border-bottom:1px solid var(--brd);white-space:nowrap}
td.num{text-align:right;font-variant-numeric:tabular-nums;color:#1e3a8a}
.row-count{padding:.35rem 1.2rem;font-size:.72rem;color:var(--mut);
  border-top:1px solid var(--brd);background:#f8fafc}
.empty{padding:1.25rem;color:var(--mut);font-style:italic}
.charts-grid{display:grid;grid-template-columns:1fr 1fr;gap:1.25rem;padding:1.25rem}
.chart-card{background:#f8fafc;border-radius:6px;padding:1rem;border:1px solid var(--brd)}
.chart-card h3{font-size:.82rem;font-weight:600;color:var(--pri-d);margin-bottom:.75rem;
  display:flex;align-items:center;gap:.45rem}
.chart-card h3::before{content:'';display:inline-block;width:3px;height:.9em;
  background:var(--pri);border-radius:2px}
.btt{position:fixed;bottom:1.25rem;right:1.25rem;background:var(--pri);color:#fff;
  border:none;border-radius:50%;width:2.25rem;height:2.25rem;font-size:1.1rem;
  cursor:pointer;box-shadow:0 2px 8px rgba(0,0,0,.25);transition:background .15s;
  display:flex;align-items:center;justify-content:center}
.btt:hover{background:var(--pri-d)}
"""

    # ── Helper: DataFrame → tabella HTML ────────────────────────────────────
    def _cell(val):
        if pd.isna(val):
            return '<td></td>'
        if isinstance(val, float):
            return f'<td class="num">{val:,.2f}</td>'
        return f'<td>{val}</td>'

    def _tbl(df):
        if df is None or df.empty:
            return '<p class="empty">Nessun dato disponibile.</p>'
        out = ['<div class="tbl-wrap"><table><thead><tr>']
        for col in df.columns:
            out.append(f'<th>{col}</th>')
        out.append('</tr></thead><tbody>')
        for _, row in df.iterrows():
            out.append('<tr>')
            out.extend(_cell(v) for v in row)
            out.append('</tr>')
        out.append('</tbody></table></div>')
        out.append(f'<p class="row-count">{len(df):,} record</p>')
        return ''.join(out)

    # ── Sezioni ─────────────────────────────────────────────────────────────
    sec_dati = f"""
<section class="sec" id="dati">
  <div class="sec-hdr"><h2>Dati</h2><span class="badge">{len(df_dati):,} righe</span></div>
  <div class="sub-sec">{_tbl(df_dati)}</div>
</section>"""

    sec_proj = f"""
<section class="sec" id="progetti">
  <div class="sec-hdr"><h2>Progetti</h2><span class="badge">{len(df_proj)} contratti</span></div>
  <div class="sub-sec">{_tbl(df_proj)}</div>
</section>"""

    sec_riep = f"""
<section class="sec" id="riepilogo">
  <div class="sec-hdr"><h2>Riepilogo Settimanale</h2></div>
  <div class="sub-sec" id="actual">
    <h3>Actual Hours (Giornate)</h3>
    {_tbl(pivot_actual)}
  </div>
  <div class="sub-sec" id="estimated">
    <h3>Estimated Hours (Giornate)</h3>
    {_tbl(pivot_estimated)}
  </div>
</section>"""

    sec_ruoli = f"""
<section class="sec" id="dettaglio-ruoli">
  <div class="sec-hdr"><h2>Dettaglio Ruoli</h2></div>
  <div class="sub-sec">{_tbl(pv_ruoli)}</div>
</section>"""

    exp_title = config.get('Export1', 'Tabella di Export')
    sec_exp = f"""
<section class="sec" id="export">
  <div class="sec-hdr"><h2>Tabella di Export</h2></div>
  <div class="sub-sec">
    <h3>{exp_title}</h3>
    {_tbl(df_exp)}
  </div>
</section>"""

    sec_grafici = """
<section class="sec" id="grafici">
  <div class="sec-hdr"><h2>Grafici</h2></div>
  <div class="charts-grid">
    <div class="chart-card">
      <h3>Actual vs Estimated per Progetto (giornate)</h3>
      <canvas id="chart-ae"></canvas>
    </div>
    <div class="chart-card">
      <h3>Trend Settimanale Giornate Actual</h3>
      <canvas id="chart-trend"></canvas>
    </div>
    <div class="chart-card">
      <h3>Top Risorse per Giornate Actual</h3>
      <canvas id="chart-risorse"></canvas>
    </div>
    <div class="chart-card">
      <h3>Stato Contratti: Riscattati vs Utilizzati</h3>
      <canvas id="chart-contratti"></canvas>
    </div>
  </div>
</section>"""

    sec_grafici_rh = """
<section class="sec" id="grafici-rh">
  <div class="sec-hdr"><h2>Grafici RH</h2></div>
  <div class="sub-sec">
    <div class="chart-card">
      <h3>Giorni Rimasti per Contratto nel Tempo</h3>
      <canvas id="chart-rh-rim"></canvas>
    </div>
  </div>
</section>"""

    sec_grafici_intesa = """
<section class="sec" id="grafici-intesa">
  <div class="sec-hdr"><h2>Grafici Intesa</h2></div>
  <div class="sub-sec">
    <div class="chart-card">
      <h3>Giorni Rimasti per Voce nel Tempo</h3>
      <canvas id="chart-intesa-rim"></canvas>
    </div>
  </div>
</section>"""

    # ── Navigazione laterale ─────────────────────────────────────────────────
    nav = f"""<nav class="sidebar">
  <div class="sb-title">Report Risorse</div>
  <div class="sb-meta">Generato il<br>{now_str}</div>
  <ul class="nav-list">
    <li class="nav-sep">Indice</li>
    <li><a href="#grafici-rh">Grafici RH</a></li>
    <li><a href="#grafici-intesa">Grafici Intesa</a></li>
    <li><a href="#dati">Dati</a></li>
    <li><a href="#progetti">Progetti</a></li>
    <li><a href="#riepilogo">Riepilogo Settimanale</a></li>
    <li><a href="#actual" class="sub">&#x2937; Actual</a></li>
    <li><a href="#estimated" class="sub">&#x2937; Estimated</a></li>
    <li><a href="#dettaglio-ruoli">Dettaglio Ruoli</a></li>
    <li><a href="#export">Tabella Export</a></li>
    <li><a href="#grafici">Grafici</a></li>
  </ul>
</nav>"""

    # ── Script Chart.js (stringa normale, non f-string, per evitare escaping) ─
    js_data = f"const D = {chart_json};"
    chart_script = (
        '<script src="https://cdn.jsdelivr.net/npm/chart.js@4/dist/chart.umd.min.js"></script>\n'
        '<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-annotation@3/dist/chartjs-plugin-annotation.min.js"></script>\n'
        '<script>\n'
        '(function(){\n'
        + js_data + '\n'
        'const PAL=["#1a56db","#f59e0b","#10b981","#ef4444","#8b5cf6","#06b6d4","#f97316"];\n'
        '\n'
        'new Chart(document.getElementById("chart-ae"),{\n'
        '  type:"bar",\n'
        '  data:{labels:D.ae.labels,datasets:[\n'
        '    {label:"Actual",    data:D.ae.actual,    backgroundColor:"#1a56db"},\n'
        '    {label:"Estimated", data:D.ae.estimated, backgroundColor:"#f59e0b"}\n'
        '  ]},\n'
        '  options:{responsive:true,plugins:{legend:{position:"top"}},\n'
        '    scales:{x:{ticks:{maxRotation:45}},y:{beginAtZero:true}}}\n'
        '});\n'
        '\n'
        'new Chart(document.getElementById("chart-trend"),{\n'
        '  type:"line",\n'
        '  data:{labels:D.trend.labels,datasets:[{\n'
        '    label:"Giornate Actual",data:D.trend.values,\n'
        '    borderColor:"#1a56db",backgroundColor:"rgba(26,86,219,0.12)",\n'
        '    fill:true,tension:0.3,pointRadius:3\n'
        '  }]},\n'
        '  options:{responsive:true,plugins:{legend:{display:false}},\n'
        '    scales:{x:{ticks:{maxRotation:45}},y:{beginAtZero:true}}}\n'
        '});\n'
        '\n'
        'new Chart(document.getElementById("chart-risorse"),{\n'
        '  type:"bar",\n'
        '  data:{labels:D.risorse.labels,datasets:[{\n'
        '    label:"Giornate",data:D.risorse.values,\n'
        '    backgroundColor:D.risorse.labels.map(function(_,i){return PAL[i%PAL.length];})\n'
        '  }]},\n'
        '  options:{indexAxis:"y",responsive:true,plugins:{legend:{display:false}},\n'
        '    scales:{x:{beginAtZero:true}}}\n'
        '});\n'
        '\n'
        'new Chart(document.getElementById("chart-contratti"),{\n'
        '  type:"bar",\n'
        '  data:{labels:D.contratti.labels,datasets:[\n'
        '    {label:"Risc. PM",    data:D.contratti.risc_pm,    backgroundColor:"#1a56db",stack:"r"},\n'
        '    {label:"Risc. Cons.", data:D.contratti.risc_cons,  backgroundColor:"#93c5fd",stack:"r"},\n'
        '    {label:"Usati PM",    data:D.contratti.usati_pm,   backgroundColor:"#f59e0b",stack:"u"},\n'
        '    {label:"Usati Cons.", data:D.contratti.usati_cons, backgroundColor:"#fcd34d",stack:"u"}\n'
        '  ]},\n'
        '  options:{responsive:true,plugins:{legend:{position:"top"}},\n'
        '    scales:{x:{ticks:{maxRotation:45}},y:{beginAtZero:true,stacked:true}}}\n'
        '});\n'
        '\n'
        'const LINE_PAL=["#1a56db","#f59e0b","#10b981","#ef4444","#8b5cf6","#06b6d4","#f97316","#84cc16","#ec4899","#78716c"];\n'
        'function makeAnnotations(){\n'
        '  var a={};\n'
        '  if(D.meta.current_week_idx>=0){\n'
        '    a.cw={type:"line",scaleID:"x",value:D.meta.current_week_idx,\n'
        '      borderColor:"rgb(220,38,38)",borderWidth:2};\n'
        '  }\n'
        '  D.meta.month_end_positions.forEach(function(p,i){\n'
        '    a["me"+i]={type:"line",scaleID:"x",value:p,\n'
        '      borderColor:"rgb(0,0,0)",borderWidth:1};\n'
        '  });\n'
        '  return a;\n'
        '}\n'
        '\n'
        'new Chart(document.getElementById("chart-rh-rim"),{\n'
        '  type:"line",\n'
        '  data:{\n'
        '    labels:D.rh.weeks,\n'
        '    datasets:D.rh.series.map(function(s,i){\n'
        '      return {label:s.label,data:s.data,\n'
        '        borderColor:LINE_PAL[i%LINE_PAL.length],\n'
        '        backgroundColor:"transparent",tension:0.3,pointRadius:3,fill:false};\n'
        '    })\n'
        '  },\n'
        '  options:{responsive:true,\n'
        '    plugins:{legend:{position:"top"},annotation:{annotations:makeAnnotations()}},\n'
        '    scales:{x:{ticks:{maxRotation:45}},y:{beginAtZero:false}}}\n'
        '});\n'
        '\n'
        'new Chart(document.getElementById("chart-intesa-rim"),{\n'
        '  type:"line",\n'
        '  data:{\n'
        '    labels:D.intesa.weeks,\n'
        '    datasets:D.intesa.series.map(function(s,i){\n'
        '      return {label:s.label,data:s.data,\n'
        '        borderColor:LINE_PAL[i%LINE_PAL.length],\n'
        '        backgroundColor:"transparent",tension:0.3,pointRadius:3,fill:false};\n'
        '    })\n'
        '  },\n'
        '  options:{responsive:true,\n'
        '    plugins:{legend:{position:"top"},annotation:{annotations:makeAnnotations()}},\n'
        '    scales:{x:{ticks:{maxRotation:45}},y:{beginAtZero:false}}}\n'
        '});\n'
        '})();\n'
        '</script>'
    )

    # ── Assemblaggio finale ──────────────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html lang="it">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>{titolo}</title>
  <style>{css}</style>
</head>
<body>
{nav}
<main>
  <div class="pg-hdr">
    <h1>{titolo}</h1>
    <p>Generato il {now_str}</p>
  </div>
{sec_grafici_rh}
{sec_grafici_intesa}
{sec_dati}
{sec_proj}
{sec_riep}
{sec_ruoli}
{sec_exp}
{sec_grafici}
</main>
<button class="btt" onclick="window.scrollTo({{top:0,behavior:'smooth'}})" title="Torna in cima">&#8679;</button>
{chart_script}
</body>
</html>"""

    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)
    log.info("HTML generato: %s", html_path)



# --- FUNZIONE PRINCIPALE ---

def elabora_dati(file_excel_input, file_cust_config, file_output, cliente_filter=''):
    """Orchestratore principale: legge input, calcola, scrive e formatta l'output.

    Flusso:
    1. Carica script.config (fisso) + cust.config e indice contratti
    2. Legge il file Excel e prepara i DataFrame
    3. (opzionale) Filtra per cliente
    4. Calcola le pivot table
    5. Prepara le righe del foglio progetti
    6. Scrive i fogli base con pandas ExcelWriter
    7. Riapre il file con openpyxl e applica formattazione
    8. Salva il file finale

    Args:
        file_excel_input: percorso del file Excel sorgente.
        file_cust_config: config specifica del cliente (contratti, Export, contatti).
        file_output:      percorso del file Excel di output da generare.
        cliente_filter:   se non vuoto, filtra le righe dove la colonna "Cliente"
                          corrisponde a questo valore (confronto case-insensitive).
    """
    try:
        config = {**carica_config('script.config'), **carica_config(file_cust_config)}
        contratti_idx = indicizza_contratti(config)

        start_w = int(config.get('StartWeek', 0))
        end_w = int(config.get('EndWeek', 99))
        weeks_limit_active = config.get('WeeksLimit', 'no').lower() == 'yes'

        oggi = datetime.now()
        anno_corrente = oggi.year
        settimana_corrente = oggi.isocalendar()[1]
        current_week_str = f"CY{anno_corrente}-W{settimana_corrente:02d}"

        if not os.path.exists(file_excel_input):
            log.error("ERRORE: %s non trovato.", file_excel_input)
            return

        log.info("1. Caricamento e preparazione dati...")
        df_src, df_dati_comp, col_proj, col_period, col_role_name, col_estimated, col_actual = carica_dati(file_excel_input, config)

        if cliente_filter:
            mask = df_dati_comp['Cliente'].str.strip().str.lower() == cliente_filter.strip().lower()
            df_dati_comp = df_dati_comp[mask].reset_index(drop=True)
            df_src = df_src[mask].reset_index(drop=True)
            log.info("   Filtro cliente '%s': %d righe selezionate.", cliente_filter, len(df_dati_comp))

        df_dati_comp['sett_calc'] = df_dati_comp[col_period].apply(estrai_settimana)

        df_per_calc, pivot_actual, pivot_estimated, pivot_role_est = calcola_pivot(
            df_dati_comp, col_period, col_actual, col_estimated,
            col_proj, col_role_name, weeks_limit_active, start_w, end_w
        )

        col_rif = "Riferimento tabella 1"
        rows_progetti = prepara_righe_progetti(df_src, df_dati_comp, df_per_calc, col_proj, col_actual, config, contratti_idx)

        log.info("2. Scrittura fogli base...")
        scrivi_fogli_base(file_output, df_dati_comp, rows_progetti)

        log.info("3. Applicazione formattazione e dati mancanti...")
        wb = load_workbook(file_output)

        ws_d = wb['dati']

        try:
            git_date = subprocess.check_output(
                ['git', 'log', '--format=%cd %h', '--date=format:%Y-%m-%d', '-1'],
                cwd=os.path.dirname(os.path.abspath(__file__)),
                stderr=subprocess.DEVNULL
            ).decode().strip()
        except Exception:
            git_date = ""
        # startrow=0 → header a riga 1, dati da riga 2 fino a riga 1+n
        ultima_riga_dati = 1 + len(df_dati_comp)
        versione_row = ultima_riga_dati + 2
        ws_d.cell(row=versione_row, column=1).value = f"Versione script: {git_date}"

        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        red_thick = Side(style='thick', color='FF0000')
        green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        fill_verde = PatternFill(fill_type="solid", fgColor="006400")
        fill_nero = PatternFill(fill_type="solid", fgColor="000000")
        font_bianco_bold = Font(color="FFFFFF", bold=True)

        formatta_tab_progetti(wb['progetti'], config, rows_progetti, weeks_limit_active, bold, center)
        # Colonne K e L del sorgente contengono lo stato di schedulazione e commit/exclude
        col_status_k = df_dati_comp.columns[10]
        col_status_l = df_dati_comp.columns[11]
        formatta_riepilogo(wb['Riepilogo Settimanale'], wb['Dettaglio Ruoli'],
                           pivot_actual, pivot_estimated, pivot_role_est,
                           current_week_str, bold, center, green_fill, red_thick,
                           df_dati_comp, col_proj, col_role_name, col_period, col_status_k, col_status_l,
                           df_per_calc, col_actual)

        riga_export = formatta_tab_export(
            wb['Tabella di Export'], config, df_per_calc, col_rif, col_actual,
            fill_verde, fill_nero, font_bianco_bold, center, right_align
        )

        if weeks_limit_active:
            aggiungi_note(wb['progetti'], wb['Tabella di Export'], anno_corrente,
                          start_w, end_w, rows_progetti, riga_export, bold)

        log.info("4. Generazione grafici Excel...")
        crea_grafici_rh(wb, rows_progetti, df_per_calc, col_proj, col_period, col_actual, bold, center)
        crea_grafici_intesa(wb, df_per_calc, col_period, col_actual, col_rif, config, bold, center)

        log.info("5. Generazione file HTML...")
        genera_html(df_dati_comp, rows_progetti, pivot_actual, pivot_estimated,
                    pivot_role_est, df_per_calc, config, col_rif, col_actual,
                    col_proj, col_period, col_estimated, file_output)

        for sheet_name in ['dati', 'progetti', 'Riepilogo Settimanale',
                            'Dettaglio Ruoli', 'Tabella di Export']:
            autofit_columns(wb[sheet_name])

        wb.save(file_output)
        log.info("SUCCESSO: File generato con tutte le intestazioni e dati Export.")

    except Exception as e:
        log.error("ERRORE: %s", str(e))
        traceback.print_exc()


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(
        description='Report settimanale risorse Red Hat Italy')
    ap.add_argument('cliente',     nargs='?', default='',
                    help='Filtro colonna Cliente (case-insensitive; vuoto = tutti)')
    ap.add_argument('input_file',  nargs='?', default='input.xlsx')
    ap.add_argument('cust_config', nargs='?', default='cust.config')
    ap.add_argument('output_file', nargs='?', default='output_elaborato.xlsx')
    args = ap.parse_args()
    elabora_dati(args.input_file, args.cust_config, args.output_file,
                 cliente_filter=args.cliente)
